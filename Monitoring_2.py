from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer
import os
import time
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl.utils.exceptions import SheetTitleException
import re
#  import shutil, os, psutil
def autostart():
    import shutil, os, psutil

    # Get current user
    user = os.getlogin()

    # Path to startup folder
    startup_path = r"C:\Users\Usuario\AppData\Roaming\Microsoft\Windows\Start Menu\Programs\Startup"

    # Path to the executable (in dist folder created by PyInstaller)
    exe_path = "dist/Definitve_Permanent_Monitoring_SIGCOM.exe"

    # Target path in startup folder
    target_path = os.path.join(startup_path, "SIGCOM_Monitor.exe")

    try:
        # Copy instead of move
        shutil.copy2(exe_path, target_path)
        print(f"Successfully added to startup: {target_path}")
    except Exception as error:
        print(f"Error adding to startup: {error}")


def configurar_hoja_activa(root_dir):
    for foldername, _, filenames in os.walk(root_dir):
        if "DEVENGADO.xlsx" in filenames:
            nombre_carpeta = os.path.basename(foldername).upper()
            archivo_excel = os.path.join(foldername, "DEVENGADO.xlsx")

            try:
                wb = openpyxl.load_workbook(archivo_excel, keep_vba=True)

                if nombre_carpeta in wb.sheetnames:
                    # Configurar la hoja deseada como activa
                    wb.active = wb[nombre_carpeta]
                    wb.save(archivo_excel)
                    print(f"Configurada hoja activa: {archivo_excel} -> '{nombre_carpeta}'")
                else:
                    print(f"Advertencia: No existe la hoja '{nombre_carpeta}' en {archivo_excel}")

                wb.close()

            except SheetTitleException:
                print(f"Error: Nombre de hoja inválido en {archivo_excel}")
            except Exception as e:
                print(f"Error procesando {archivo_excel}: {str(e)}")
directorio_raiz = "RUTA/A/TUS/CARPETAS"
configurar_hoja_activa(directorio_raiz)

base_path = r"\\10.5.130.24\Abastecimiento\Compartido Abastecimiento\Otros\SIGCOM"
work_directory = r'C:\Users\Usuario\Downloads'  # Where data files are located
meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']


class FileMonitorHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        print(f'File created: {event.src_path}')
        self.process_file(event.src_path)

    def on_moved(self, event):
        if event.is_directory:
            return
        print(f'File renamed: {event.src_path} -> {event.dest_path}')
        self.process_file(event.dest_path)

    def on_modified(self, event):
        if event.is_directory:
            # Check if this is a directory we're monitoring
            verificar_carpetas(event.src_path)
            return
        print(f'File modified: {event.src_path}')
        self.process_file(event.src_path)

    def check_required_files_exist(self):
        """Check if Codigos_Clasificador_Compilado.xlsx exists in work_directory."""
        original_dir = os.getcwd()
        try:
            os.chdir(work_directory)
            if not os.path.exists("Codigos_Clasificador_Compilado.xlsx"):
                print(f"ERROR: Codigos_Clasificador_Compilado.xlsx is missing in {work_directory}")
                return False
            return True
        finally:
            os.chdir(original_dir)

    def check_devengado_exists(self, folder_path):
        """Check if DEVENGADO file exists in the same folder as the BASE DISTRIBUCION file."""
        if not os.path.exists(folder_path):
            return False

        devengado_files = [f for f in os.listdir(folder_path) if f.startswith("DEVENGADO") and f.endswith(".xlsx")]
        if not devengado_files:
            print(f"ERROR: No DEVENGADO file found in {folder_path}")
            return False

        print(f"Found DEVENGADO file: {devengado_files[0]} in {folder_path}")
        return True

    def check_modified_exists(self, file_path):
        """Check if a modified version of the file already exists."""
        folder = os.path.dirname(file_path)
        filename = os.path.basename(file_path)
        modified_file = os.path.join(folder, f"Modified_{filename}")

        if os.path.exists(modified_file):
            print(f"WARNING: Modified file already exists: {modified_file}")
            print("Processing skipped to avoid overwriting existing modified file.")
            return True
        return False

    def process_file(self, file_path):
        # Only process Excel files starting with "BASE DISTRIBUCION GASTO GENERAL"
        filename = os.path.basename(file_path)
        folder_path = os.path.dirname(file_path)

        # Skip temporary Excel files or already modified files
        if (filename.startswith('~$') or
                filename.startswith('Modified_') or
                not filename.endswith('.xlsx') or
                not filename.startswith('BASE DISTRIBUCION GASTO GENERAL')):
            return

        # Check if modified file already exists
        if self.check_modified_exists(file_path):
            return

        # Check if required files exist before processing
        if not self.check_required_files_exist():
            print("Processing aborted due to missing Codigos_Clasificador_Compilado.xlsx.")
            return

        # Check if DEVENGADO file exists in the same folder
        if not self.check_devengado_exists(folder_path):
            print("Processing aborted due to missing DEVENGADO file in the same folder.")
            return

        try:
            print(f"Processing Excel file: {file_path}")
            #update_excel_with_xlwings(file_path)
            print(f"Processing complete. Modified file saved.")
        except Exception as e:
            print(f"Error processing file: {str(e)}")

def configurar_hoja_activa(root_dir):
    for foldername, _, filenames in os.walk(root_dir):
        if "DEVENGADO.xlsx" in filenames:
            nombre_carpeta = os.path.basename(foldername).upper()
            archivo_excel = os.path.join(foldername, "DEVENGADO.xlsx")

            try:
                wb = openpyxl.load_workbook(archivo_excel, keep_vba=True)

                if nombre_carpeta in wb.sheetnames:
                    # Configurar la hoja deseada como activa
                    wb.active = wb[nombre_carpeta]
                    wb.save(archivo_excel)
                    print(f"Configurada hoja activa: {archivo_excel} -> '{nombre_carpeta}'")
                else:
                    print(f"Advertencia: No existe la hoja '{nombre_carpeta}' en {archivo_excel}")

                wb.close()

            except SheetTitleException:
                print(f"Error: Nombre de hoja inválido en {archivo_excel}")
            except Exception as e:
                print(f"Error procesando {archivo_excel}: {str(e)}")

def verificar_carpetas(carpeta_modificada=None):
    """Verifica carpetas de años/meses y busca archivos Excel para procesar."""
    for año in range(2024, 2041):
        for mes in meses:
            carpeta_mes = os.path.join(base_path, str(año), mes)

            # Si se especificó una carpeta, solo procesar esa
            if carpeta_modificada and carpeta_modificada != carpeta_mes:
                continue

            if not os.path.exists(carpeta_mes):
                continue

            # Buscar archivos Excel que coincidan con el patrón
            archivos = os.listdir(carpeta_mes)
            base_files = [a for a in archivos if a.endswith('.xlsx') and a.startswith('BASE DISTRIBUCION GASTO GENERAL')]

            # Check if there's at least one DEVENGADO file in the folder
            devengado_files = [a for a in archivos if a.startswith("DEVENGADO") and a.endswith(".xlsx")]
            if not devengado_files or not base_files:
                continue

            for archivo in base_files:
                file_path = os.path.join(carpeta_mes, archivo)

                # Create an instance of the handler to use its methods
                handler = FileMonitorHandler()

                # Check if modified file already exists
                if handler.check_modified_exists(file_path):
                    continue

                # Verify required files exist
                if not handler.check_required_files_exist():
                    print("Codigos_Clasificador_Compilado.xlsx missing, cannot process.")
                    continue

                try:
                    print(f"Processing Excel file: {file_path}")
                    #update_excel_with_xlwings(file_path)
                    print(f"Processing complete. Modified file saved.")
                except Exception as e:
                    print(f"Error processing file: {str(e)}")